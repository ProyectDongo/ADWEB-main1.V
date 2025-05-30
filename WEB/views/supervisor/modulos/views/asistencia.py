from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from WEB.models import RegistroEmpresas, VigenciaPlan,Usuario
from WEB.forms import UsuarioForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,View
from django.views import View
from django.urls import reverse,reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from WEB.models import RegistroEmpresas, Usuario, VigenciaPlan, Horario, Turno ,RegistroEntrada,DiaHabilitado,Notificacion, Ubicacion,SeguroCesantia,PerfilUsuario,ContactoUsuario,InformacionAdicional,InformacionBancaria,InformacionComplementaria,Prevision,Otros,AntecedentesConducir,ExamenesMutual,GrupoFamiliar,Capacitacion,LicenciasMedicas,NivelEstudios
from WEB.forms import UsuarioForm, HorarioForm, TurnoForm ,RegistroEntradaForm, PerfilUsuarioForm, ContactoUsuarioForm, InformacionBancariaForm,InformacionAdicionalForm, SeguroCesantiaForm, PrevisionForm, OtrosForm,AntecedentesConducirForm, NivelEstudiosForm, InformacionComplementariaForm,ExamenesMutualFormSet, GrupoFamiliarFormSet, CapacitacionFormSet, LicenciasMedicasFormSet, AntecedentesConducirFormSet
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta, date
from calendar import monthrange
import logging
from django import template
from django.db import DatabaseError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter





logger = logging.getLogger(__name__)
# Vista para el home del supervisor en el módulo de asistencia
@login_required
def supervisor_home_asistencia(request, empresa_id, vigencia_plan_id):
    if not request.user.is_authenticated:
        return render(request, 'error/error.html', {'message': 'Debes iniciar sesión para acceder a esta página'})
    if request.user.role != 'supervisor':
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})

    empresa = get_object_or_404(RegistroEmpresas, id=empresa_id)
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id, empresa=empresa)

    if request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'No tienes permiso para este módulo'})

    usuarios = Usuario.objects.filter(vigencia_plan=vigencia_plan).select_related('vigencia_plan').order_by('role', 'last_name')
    supervisores = usuarios.filter(role='supervisor')
    trabajadores = usuarios.filter(role='trabajador')

    context = {
        'empresa': empresa,
        'vigencia_plan': vigencia_plan,
        'supervisores': supervisores,
        'trabajadores': trabajadores,
        'form': UsuarioForm()
    }
    return render(request, 'home/supervisores/supervisor_home_asistencia.html', context)





# nueva vista para manejar las notificaciones del supervisor
@login_required
def notificaciones_supervisor_json(request, vigencia_plan_id):
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.role != 'supervisor' or request.user.vigencia_plan != vigencia_plan:
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    try:
        notificaciones = Notificacion.objects.filter(
            worker__vigencia_plan=vigencia_plan,
            leido=False
        ).select_related('worker').order_by('-timestamp')[:5]

        data = {
            'count': notificaciones.count(),
            'notifications': [
                {
                    'id': n.id,
                    'tipo': n.tipo,
                    'timestamp': n.timestamp.isoformat(),
                    'worker': n.worker.get_full_name(),
                    'worker_id': n.worker.id,
                    'ip_address': str(n.ip_address) if n.ip_address else None,
                    'ubicacion_nombre': (
                        Ubicacion.objects.filter(vigencia_plan=vigencia_plan, ip_address=n.ip_address)
                        .first().nombre if Ubicacion.objects.filter(vigencia_plan=vigencia_plan, ip_address=n.ip_address).exists()
                        else None
                    ),
                } for n in notificaciones
            ]
        }
        return JsonResponse(data)
    except DatabaseError as e:
        logger.error(f"Error de base de datos: {e}")
        return JsonResponse({'error': 'Error al acceder a la base de datos'}, status=500)





# nueva vista para establecer el nombre de la ubicación
@login_required
def set_ubicacion_nombre(request, vigencia_plan_id, ip_address):
    if request.user.role != 'supervisor':
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.vigencia_plan != vigencia_plan:
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            ubicacion, created = Ubicacion.objects.get_or_create(
                vigencia_plan=vigencia_plan,
                ip_address=ip_address,
                defaults={'nombre': nombre}
            )
            if not created:
                ubicacion.nombre = nombre
                ubicacion.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def registros_entrada_vigencia(request, vigencia_plan_id):
    if request.user.role != 'supervisor':
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})
    
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'No tienes permiso para este módulo'})
    
    registros = RegistroEntrada.objects.filter(
        trabajador__vigencia_plan=vigencia_plan
    ).select_related('trabajador').order_by('-hora_entrada')
    
    empresa_id = vigencia_plan.empresa.id
    
    context = {
        'vigencia_plan': vigencia_plan,
        'registros': registros,
        'empresa_id': empresa_id,
        'vigencia_plan_id': vigencia_plan_id,
    }
    return render(request, 'home/supervisores/registros_entrada.html', context)






class ValidationView(View):
    def get(self, request):
        rut = request.GET.get('rut')
        email = request.GET.get('email')
        
        if rut:
            exists = Usuario.objects.filter(rut=rut).exists()
            return JsonResponse({'exists': exists})
            
        if email:
            exists = Usuario.objects.filter(email=email).exists()
            return JsonResponse({'exists': exists})
            
        return JsonResponse({'error': 'Campo inválido'}, status=400)
    
class GetFormTemplateView(View):
    def get(self, request, action):
        form = UsuarioForm()
        if action == 'create':
            return render(request, 'formularios/supervisor/supervisor.asistencia.html', {'form': form})
        elif action == 'edit':
            form.fields['password'].required = False  # Hacer opcional el campo password
            return render(request, 'formularios/supervisor/supervisor.edit.html', {'form': form})
        return HttpResponse(status=404)
    







# AQUI EN ADELANTE ESTA TODO LO NUEVO LA IMPLEMETNACION DE GESTION DE HORAIOS Y TURNOS Y BLOQUEOS ACCESO 

# Manejo de Horarios,creación y edición

class HorarioListView(LoginRequiredMixin, ListView):
    model = Horario
    template_name = 'home/supervisores/horario/horarios_list.html'
    context_object_name = 'horarios'

    def get_queryset(self):
        # Filtra los horarios por la empresa del usuario autenticado
        return Horario.objects.filter(empresa=self.request.user.empresa)

    def get_context_data(self, **kwargs):
        # Obtiene el contexto base
        context = super().get_context_data(**kwargs)
        # Agrega empresa_id y vigencia_plan_id al contexto
        context['empresa_id'] = self.request.user.empresa.id if self.request.user.empresa else None
        context['vigencia_plan_id'] = self.request.user.vigencia_plan.id if self.request.user.vigencia_plan else None
        return context
class HorarioCreateView(LoginRequiredMixin, CreateView):
    model = Horario
    form_class = HorarioForm
    template_name = 'home/supervisores/horario/horario_form.html'
    success_url = reverse_lazy('horarios_list')

    def form_valid(self, form):
        form.instance.empresa = self.request.user.empresa
        return super().form_valid(form)
    
class HorarioUpdateView(LoginRequiredMixin, UpdateView):
    model = Horario
    form_class = HorarioForm
    template_name = 'home/supervisores/horario/horario_form.html'
    success_url = reverse_lazy('horarios_list')

class HorarioDeleteView(LoginRequiredMixin, DeleteView):
    model = Horario
    template_name = 'home/supervisores/horario/horario_confirm_delete.html'
    success_url = reverse_lazy('horarios_list')










# Manejo de Turnos,Creación y Edición

class TurnoListView(LoginRequiredMixin, ListView):
    model = Turno
    template_name = 'home/supervisores/turno/turnos_list.html'
    context_object_name = 'turnos'

    def get_queryset(self):
        return Turno.objects.filter(empresa=self.request.user.empresa)
    
    # ESTO MANEJA EL VOLVER AL HOME
    def get_queryset(self):
        # Filtra los horarios por la empresa del usuario autenticado
        return Turno.objects.filter(empresa=self.request.user.empresa)

    def get_context_data(self, **kwargs):
        # Obtiene el contexto base
        context = super().get_context_data(**kwargs)
        # Agrega empresa_id y vigencia_plan_id al contexto
        context['empresa_id'] = self.request.user.empresa.id if self.request.user.empresa else None
        context['vigencia_plan_id'] = self.request.user.vigencia_plan.id if self.request.user.vigencia_plan else None
        return context


class TurnoCreateView(LoginRequiredMixin, CreateView):
    model = Turno
    form_class = TurnoForm
    template_name = 'home/supervisores/turno/turno_form.html'
    success_url = reverse_lazy('turnos_list')

    def form_valid(self, form):
        form.instance.empresa = self.request.user.empresa
        return super().form_valid(form)

class TurnoUpdateView(LoginRequiredMixin, UpdateView):
    model = Turno
    form_class = TurnoForm
    template_name = 'home/supervisores/turno/turno_form.html'
    success_url = reverse_lazy('turnos_list')

class TurnoDeleteView(LoginRequiredMixin, DeleteView):
    model = Turno
    template_name = 'home/supervisores/turno/turno_confirm_delete.html'
    success_url = reverse_lazy('turnos_list')









    
# Creación y edición de usuarios

class UserCreateUpdateView(LoginRequiredMixin, View):
    def get(self, request, vigencia_plan_id, user_id=None):  
        if user_id:
            user = get_object_or_404(Usuario, pk=user_id)
            data = {
                'username': user.username,
                'rut': user.rut,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'celular': user.celular, 
                'role': user.role,
                'horario': user.horario_id,
                'turno': user.turno_id,
                'metodo_registro_permitido': user.metodo_registro_permitido
            }
            return JsonResponse(data)
        return JsonResponse({'error': 'ID de usuario no proporcionado'}, status=400)

    def delete(self, request, vigencia_plan_id, user_id):
        user = get_object_or_404(Usuario, pk=user_id)
        try:
            user.delete()
            return JsonResponse({'message': 'Usuario eliminado exitosamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def post(self, request, vigencia_plan_id, user_id=None):
        vigencia_plan = get_object_or_404(VigenciaPlan, pk=vigencia_plan_id)
        user = get_object_or_404(Usuario, pk=user_id) if user_id else None
        
        form = UsuarioForm(request.POST, instance=user)
        if form.is_valid():
            if form.cleaned_data['role'] == 'admin':
                return JsonResponse({'errors': {'role': ['No se permite asignar el rol de administrador']}}, status=403)
            
            user = form.save(commit=False)
            user.empresa = vigencia_plan.empresa
            user.vigencia_plan = vigencia_plan
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()
            return JsonResponse({
                'message': 'Usuario guardado exitosamente',
                'redirect_url': reverse('supervisor_home_asistencia', kwargs={
                    'empresa_id': vigencia_plan.empresa.id,
                    'vigencia_plan_id': vigencia_plan.id
                })
            }, status=200)
        else:
            errors = {f: [str(e) for e in e_list] for f, e_list in form.errors.items()}
            return JsonResponse({'errors': errors}, status=400)
        









# Manejo de Calendario de Turnos

class CalendarioTurnoView(LoginRequiredMixin, View):
    template_name = 'home/supervisores/turno/calendario_turno.html'

    def get(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        año = int(request.GET.get('año', date.today().year))
        
        # Generar todos los días del año
        inicio_año = date(año, 1, 1)
        fin_año = date(año, 12, 31)
        dias = []
        current = inicio_año
        while current <= fin_año:
            dia_habilitado = DiaHabilitado.objects.filter(usuario=usuario, fecha=current).first()
            estado = dia_habilitado.habilitado if dia_habilitado else usuario.debe_trabajar(current)
            dias.append({
                'fecha': current,
                'habilitado': estado,
                'mes': current.month,
                'dia_semana': current.weekday(),
            })
            current += timedelta(days=1)

        # Organizar los días por mes
        meses = {}
        for i in range(1, 13):
            meses[i] = [d for d in dias if d['mes'] == i]

        context = {
            'usuario': usuario,
            'año': año,
            'meses': meses,
            'empresa_id': usuario.vigencia_plan.empresa.id,
            'vigencia_plan_id': usuario.vigencia_plan.id,
        }
        return render(request, self.template_name, context)

    def post(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        año = int(request.POST.get('año', date.today().year))
        
        # Procesar los datos enviados
        inicio_año = date(año, 1, 1)
        fin_año = date(año, 12, 31)
        current = inicio_año
        while current <= fin_año:
            fecha_str = current.strftime('%Y-%m-%d')
            habilitado = request.POST.get(f'dia_{fecha_str}') == 'on'
            DiaHabilitado.objects.update_or_create(
                usuario=usuario,
                fecha=current,
                defaults={'habilitado': habilitado}
            )
            current += timedelta(days=1)

        # Agregar mensaje de éxito
        messages.success(request, f'Los cambios para {usuario.get_full_name()} ({año}) se han guardado correctamente.')

        # Redirigir a supervisor_home_asistencia con empresa_id y vigencia_plan_id
        return redirect('supervisor_home_asistencia', 
                        empresa_id=usuario.vigencia_plan.empresa.id, 
                        vigencia_plan_id=usuario.vigencia_plan.id)









# Actualizar el estado de un día específico    

class ActualizarDiaView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        # Obtener el usuario o devolver 404 si no existe
        usuario = get_object_or_404(Usuario, pk=user_id)
        
        # Obtener datos del cuerpo de la solicitud
        fecha = request.POST.get('fecha')
        habilitado = request.POST.get('habilitado') == 'true'
        
        # Validar que se proporcione la fecha
        if not fecha:
            return JsonResponse({'success': False, 'error': 'Fecha no proporcionada'}, status=400)
        
        # Actualizar o crear el registro en DiaHabilitado
        dia, created = DiaHabilitado.objects.update_or_create(
            usuario=usuario,
            fecha=fecha,
            defaults={'habilitado': habilitado}
        )
        
        # Devolver respuesta JSON exitosa
        return JsonResponse({'success': True})
    

@login_required
def user_full_info(request, user_id):
    user = get_object_or_404(Usuario, id=user_id)
    
    # Verificación de permisos
    if request.user.role not in ['supervisor', 'admin'] or user.vigencia_plan != request.user.vigencia_plan:
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})

    # Obtener o crear instancias de modelos OneToOne
    perfil, _ = PerfilUsuario.objects.get_or_create(usuario=user)
    contacto, _ = ContactoUsuario.objects.get_or_create(usuario=user)
    info_bancaria, _ = InformacionBancaria.objects.get_or_create(usuario=user)
    info_adicional, _ = InformacionAdicional.objects.get_or_create(usuario=user)
    seguro_cesantia, _ = SeguroCesantia.objects.get_or_create(usuario=user)
    prevision, _ = Prevision.objects.get_or_create(usuario=user)
    otros, _ = Otros.objects.get_or_create(usuario=user)
    nivel_estudios, _ = NivelEstudios.objects.get_or_create(usuario=user)
    info_complementaria, _ = InformacionComplementaria.objects.get_or_create(usuario=user)

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        # Manejo de formularios OneToOne
        if form_type == 'usuario':
            form = UsuarioForm(request.POST, instance=user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información básica actualizada.')
        
        elif form_type == 'perfil':
            form = PerfilUsuarioForm(request.POST, instance=perfil)
            if form.is_valid():
                form.save()
                messages.success(request, 'Perfil actualizado.')
        
        elif form_type == 'contacto':
            form = ContactoUsuarioForm(request.POST, instance=contacto)
            if form.is_valid():
                form.save()
                messages.success(request, 'Contacto actualizado.')
        
        elif form_type == 'info_bancaria':
            form = InformacionBancariaForm(request.POST, instance=info_bancaria)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información bancaria actualizada.')
        
        elif form_type == 'info_adicional':
            form = InformacionAdicionalForm(request.POST, instance=info_adicional)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información adicional actualizada.')
        
        elif form_type == 'seguro_cesantia':
            form = SeguroCesantiaForm(request.POST, instance=seguro_cesantia)
            if form.is_valid():
                form.save()
                messages.success(request, 'Seguro de cesantía actualizado.')
        
        elif form_type == 'prevision':
            form = PrevisionForm(request.POST, instance=prevision)
            if form.is_valid():
                form.save()
                messages.success(request, 'Previsión actualizada.')
        
        elif form_type == 'otros':
            form = OtrosForm(request.POST, instance=otros)
            if form.is_valid():
                form.save()
                messages.success(request, 'Otros datos actualizados.')
        
        elif form_type == 'nivel_estudios':
            form = NivelEstudiosForm(request.POST, instance=nivel_estudios)
            if form.is_valid():
                form.save()
                messages.success(request, 'Nivel de estudios actualizado.')
        
        elif form_type == 'info_complementaria':
            form = InformacionComplementariaForm(request.POST, instance=info_complementaria)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información complementaria actualizada.')
        
        # Manejo de formsets ForeignKey
        elif form_type == 'examenes':
            formset = ExamenesMutualFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Exámenes mutual actualizados.')
            if 'export_examenes' in request.POST:
                queryset = user.examenes_mutual.all()
                fields = ['tipo_examen', 'fecha_examen', 'fecha_vencimiento']
                return export_to_excel(request, queryset, fields, f"examenes_mutual_{user.username}")
        
        elif form_type == 'grupo_familiar':
            formset = GrupoFamiliarFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Grupo familiar actualizado.')
            if 'export_grupo_familiar' in request.POST:
                queryset = user.grupo_familiar.all()
                fields = ['rut_carga', 'nombre_carga', 'fecha_nacimiento', 'edad', 'sexo']
                return export_to_excel(request, queryset, fields, f"grupo_familiar_{user.username}")
        
        elif form_type == 'capacitacion':
            formset = CapacitacionFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Capacitaciones actualizadas.')
            if 'export_capacitaciones' in request.POST:
                queryset = user.capacitaciones.all()
                fields = ['descripcion', 'horas', 'institucion']
                return export_to_excel(request, queryset, fields, f"capacitaciones_{user.username}")
        
        elif form_type == 'licencias_medicas':
            formset = LicenciasMedicasFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Licencias médicas actualizadas.')
            if 'export_licencias_medicas' in request.POST:
                queryset = user.licencias_medicas.all()
                fields = ['tipo_accidente', 'clasificacion_accidente', 'fecha_inicio_reposo', 
                          'fecha_termino', 'fecha_alta', 'dias_reposo']
                return export_to_excel(request, queryset, fields, f"licencias_medicas_{user.username}")
        
        elif form_type == 'antecedentes_conducir':
            formset = AntecedentesConducirFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Antecedentes de conducir actualizados.')
            if 'export_antecedentes_conducir' in request.POST:
                queryset = user.antecedentes_conducir.all()
                fields = ['tipo_licencia', 'municipalidad', 'fecha_ultimo_control', 
                          'fecha_vencimiento', 'hoja_vida_conducir']
                return export_to_excel(request, queryset, fields, f"antecedentes_conducir_{user.username}")
        
        return redirect('user_full_info', user_id=user.id)

    # GET: Preparar formularios y formsets
    usuario_form = UsuarioForm(instance=user)
    perfil_form = PerfilUsuarioForm(instance=perfil)
    contacto_form = ContactoUsuarioForm(instance=contacto)
    info_bancaria_form = InformacionBancariaForm(instance=info_bancaria)
    info_adicional_form = InformacionAdicionalForm(instance=info_adicional)
    seguro_cesantia_form = SeguroCesantiaForm(instance=seguro_cesantia)
    prevision_form = PrevisionForm(instance=prevision)
    otros_form = OtrosForm(instance=otros)
    nivel_estudios_form = NivelEstudiosForm(instance=nivel_estudios)
    info_complementaria_form = InformacionComplementariaForm(instance=info_complementaria)
    
    examenes_formset = ExamenesMutualFormSet(instance=user)
    grupo_familiar_formset = GrupoFamiliarFormSet(instance=user)
    capacitacion_formset = CapacitacionFormSet(instance=user)
    licencias_medicas_formset = LicenciasMedicasFormSet(instance=user)
    antecedentes_conducir_formset = AntecedentesConducirFormSet(instance=user)

    context = {
        'user': user,
        'usuario_form': usuario_form,
        'perfil_form': perfil_form,
        'contacto_form': contacto_form,
        'info_bancaria_form': info_bancaria_form,
        'info_adicional_form': info_adicional_form,
        'seguro_cesantia_form': seguro_cesantia_form,
        'prevision_form': prevision_form,
        'otros_form': otros_form,
        'nivel_estudios_form': nivel_estudios_form,
        'info_complementaria_form': info_complementaria_form,
        'examenes_formset': examenes_formset,
        'grupo_familiar_formset': grupo_familiar_formset,
        'capacitacion_formset': capacitacion_formset,
        'licencias_medicas_formset': licencias_medicas_formset,
        'antecedentes_conducir_formset': antecedentes_conducir_formset,
    }
    
    return render(request, 'home/supervisores/DatosCompletos/user_full_info.html', context)

# Función para exportar a Excel
def export_to_excel(request, queryset, fields, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Escribir encabezados
    for col_num, field in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = field

    # Escribir datos
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            value = getattr(obj, field, '')
            if isinstance(value, (date, datetime)):
                value = value.strftime("%Y-%m-%d")
            elif value is None:
                value = ''
            ws[f"{col_letter}{row_num}"] = str(value)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    wb.save(response)
    return response