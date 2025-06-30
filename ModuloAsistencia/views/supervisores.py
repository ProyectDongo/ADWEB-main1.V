from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,View
from django.views import View
from django.urls import reverse,reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from WEB.models import *
from WEB.forms.modulo_asistencia.forms import *
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta, date
from calendar import monthrange
import logging
from django import template
from django.db import DatabaseError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from zipfile import ZipFile
from django.conf import settings
import json
import random
import string
from ModuloAsistencia.models import RegistroEntrada
from WEB.views.scripts.utils import hay_pagos_atrasados
from django.views.decorators.http import require_POST
from django.core.mail import send_mail






def get_today_assignment(user):
    today = timezone.now().date()
    return AsignacionDiaria.objects.filter(usuario=user, fecha=today).first()

def generate_access_code(user):
    code = ''.join(random.choices(string.digits, k=6))
    expires_at = timezone.now() + timedelta(minutes=5)
    AccessCode.objects.create(user=user, code=code, expires_at=expires_at)
    return code

def send_access_code_email(user, code):
    subject = "Código de acceso para entrada tardía"
    message = f"Tu código de acceso es: {code}\nEste código expira en 5 minutos."
    send_mail(subject, message, 'from@example.com', [user.email], fail_silently=False)



logger = logging.getLogger(__name__)
@login_required
def supervisor_home_asistencia(request, empresa_id, vigencia_plan_id):
    if not request.user.is_authenticated:
        return render(request, 'error/error.html', {'message': 'Debes iniciar sesión para acceder a esta página'})
    if request.user.role != 'supervisor':
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})

    empresa = get_object_or_404(RegistroEmpresas, id=empresa_id)
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id, empresa=empresa)

    if request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'No tienes permiso para este módulo'})

    usuarios = Usuario.objects.filter(vigencia_plan=vigencia_plan).select_related('vigencia_plan').order_by('role', 'last_name')
    supervisores = usuarios.filter(role='supervisor')
    trabajadores = usuarios.filter(role='trabajador')

    # Obtener notificaciones de llegada tarde
    notifications = LateArrivalNotification.objects.filter(
        user__vigencia_plan=vigencia_plan,
        code_sent=False
    ).select_related('user').order_by('-timestamp')

    context = {
        'empresa': empresa,
        'vigencia_plan': vigencia_plan,
        'supervisores': supervisores,
        'trabajadores': trabajadores,
        'form': UsuarioForm(),
        'notifications': notifications,
    }
    return render(request, 'Supervisores/Modulo_asistencia/home/supervisor_home_asistencia.html', context)




@login_required
def late_arrival_notifications_json(request, vigencia_plan_id):
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.role != 'supervisor' or request.user.vigencia_plan != vigencia_plan:
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    notifications = LateArrivalNotification.objects.filter(
            user__vigencia_plan=vigencia_plan,
            code_sent=False
        ).select_related('user').order_by('-timestamp')[:5]

    data = {
        'notifications': [
            {
                'id': n.id,
                'user': n.user.get_full_name(),
                'user_id': n.user.id,
                'timestamp': n.timestamp.isoformat(),
            } for n in notifications
        ]
    }
    return JsonResponse(data)

@require_POST
@login_required
def clear_notifications(request, vigencia_plan_id):
    if request.user.role != 'supervisor':
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)
    
    notifications = LateArrivalNotification.objects.filter(
        user__vigencia_plan_id=vigencia_plan_id,
        code_sent=False
    )
    notifications.update(code_sent=True)  # Marca todas como enviadas
    return JsonResponse({'success': True})





@require_POST
def send_access_code(request, notification_id):
    logger.info(f'Solicitud para enviar código, notificación ID: {notification_id}')
    notification = get_object_or_404(LateArrivalNotification, id=notification_id)
    if request.user.role != 'supervisor':
        logger.warning(f'Acceso no autorizado por usuario: {request.user}')
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)
    
    try:
        code = generate_access_code(notification.user)
        send_access_code_email(notification.user, code)
        notification.code_sent = True
        notification.save()
        logger.info(f'Código enviado para notificación ID: {notification_id}')
        return JsonResponse({
            'success': True,
            'code': code,
            'message': 'Código enviado al trabajador y registrado exitosamente'
        })
    except Exception as e:
        logger.error(f'Error al enviar código: {e}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@require_POST
@login_required
def send_code_email(request, notification_id):
    notification = get_object_or_404(LateArrivalNotification, id=notification_id)
    if request.user.role != 'supervisor':
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)
    
    code = request.POST.get('code')
    if not code:
        return JsonResponse({'error': 'Código no proporcionado'}, status=400)
    
    send_access_code_email(notification.user, code) 
    notification.code_sent = True
    notification.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Código enviado al trabajador exitosamente'
    })




@login_required
def late_arrival_history(request, vigencia_plan_id):
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.role != 'supervisor' or request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})

    notifications = LateArrivalNotification.objects.filter(
        user__vigencia_plan=vigencia_plan
    ).select_related('user').order_by('-timestamp')

    context = {
        'notifications': notifications,
        'vigencia_plan': vigencia_plan,
    }
    return render(request, 'Supervisores/Modulo_asistencia/home/retrasos.html', context)




# REPORTES DE ASISTENCIA
def generate_daily_report(user, date):
    entries = RegistroEntrada.objects.filter(
        trabajador=user,
        hora_entrada__date=date
    ).order_by('hora_entrada')
    report = f"Reporte diario para {date.strftime('%d/%m/%Y')}:\n"
    for entry in entries:
        report += f"Entrada: {entry.hora_entrada.strftime('%H:%M')}, Salida: {entry.hora_salida.strftime('%H:%M') if entry.hora_salida else 'Pendiente'}\n"
    return report



def generate_weekly_report(user, start_date):
    end_date = start_date + timedelta(days=6)
    entries = RegistroEntrada.objects.filter(
        trabajador=user,
        hora_entrada__date__range=(start_date, end_date)
    ).order_by('hora_entrada')
    report = f"Reporte semanal desde {start_date.strftime('%d/%m/%Y')} hasta {end_date.strftime('%d/%m/%Y')}:\n"
    for entry in entries:
        report += f"{entry.hora_entrada.strftime('%d/%m/%Y')}: Entrada: {entry.hora_entrada.strftime('%H:%M')}, Salida: {entry.hora_salida.strftime('%H:%M') if entry.hora_salida else 'Pendiente'}\n"
    return report


def generate_monthly_report(user, year, month):
    entries = RegistroEntrada.objects.filter(
        trabajador=user,
        hora_entrada__year=year,
        hora_entrada__month=month
    ).order_by('hora_entrada')
    report = f"Reporte mensual para {year}-{month:02d}:\n"
    for entry in entries:
        report += f"{entry.hora_entrada.strftime('%d/%m/%Y')}: Entrada: {entry.hora_entrada.strftime('%H:%M')}, Salida: {entry.hora_salida.strftime('%H:%M') if entry.hora_salida else 'Pendiente'}\n"
    return report

def send_reports():
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    users = Usuario.objects.filter(email__isnull=False)
    
    # Reporte diario
    for user in users:
        report = generate_daily_report(user, yesterday)
        send_mail(
            f"Reporte diario de asistencia - {yesterday.strftime('%d/%m/%Y')}",
            report,
            'from@example.com',
            [user.email],
            fail_silently=True
        )
    
    # Reporte semanal (si es lunes)
    if today.weekday() == 0:
        start_date = today - timedelta(days=7)
        for user in users:
            report = generate_weekly_report(user, start_date)
            send_mail(
                f"Reporte semanal de asistencia - {start_date.strftime('%d/%m/%Y')}",
                report,
                'from@example.com',
                [user.email],
                fail_silently=True
            )
    
    # Reporte mensual (si es el primer día del mes)
    if today.day == 1:
        last_month = today - timedelta(days=1)
        for user in users:
            report = generate_monthly_report(user, last_month.year, last_month.month)
            send_mail(
                f"Reporte mensual de asistencia - {last_month.strftime('%Y-%m')}",
                report,
                'from@example.com',
                [user.email],
                fail_silently=True
            )












# vista que maneja ingerso y salidas normales 
@login_required
def notificaciones_supervisor_json(request, vigencia_plan_id):
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.role != 'supervisor' or request.user.vigencia_plan != vigencia_plan:
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    try:
        notificaciones = Notificacion.objects.filter(
            worker__vigencia_plan=vigencia_plan,
            leido=False
        ).select_related('worker').order_by('-timestamp')[:5]

        data = {
            'count': notificaciones.count(),
            'notifications': [
                {
                    'id': n.id,
                    'tipo': n.tipo,
                    'timestamp': n.timestamp.isoformat(),
                    'worker': n.worker.get_full_name(),
                    'worker_id': n.worker.id,
                    'ip_address': str(n.ip_address) if n.ip_address else None,
                    'ubicacion_nombre': (
                        Ubicacion.objects.filter(vigencia_plan=vigencia_plan, ip_address=n.ip_address)
                        .first().nombre if Ubicacion.objects.filter(vigencia_plan=vigencia_plan, ip_address=n.ip_address).exists()
                        else None
                    ),
                    'latitud': n.worker.entradas.filter(hora_entrada__date=n.timestamp.date(), metodo='geo').first().latitud if n.worker.entradas.filter(hora_entrada__date=n.timestamp.date(), metodo='geo').exists() else None,
                    'longitud': n.worker.entradas.filter(hora_entrada__date=n.timestamp.date(), metodo='geo').first().longitud if n.worker.entradas.filter(hora_entrada__date=n.timestamp.date(), metodo='geo').exists() else None,
                } for n in notificaciones
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error: {e}")
        return JsonResponse({'error': 'Error al procesar notificaciones'}, status=500)









# nueva vista para establecer el nombre de la ubicación
@login_required
def set_ubicacion_nombre(request, vigencia_plan_id, ip_address):
    if request.user.role != 'supervisor':
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.vigencia_plan != vigencia_plan:
        return JsonResponse({'error': 'Acceso no autorizado'}, status=403)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            ubicacion, created = Ubicacion.objects.get_or_create(
                vigencia_plan=vigencia_plan,
                ip_address=ip_address,
                defaults={'nombre': nombre}
            )
            if not created:
                ubicacion.nombre = nombre
                ubicacion.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
    return JsonResponse({'error': 'Método no permitido'}, status=405)








@login_required
def ver_mapa_registros(request, vigencia_plan_id):
    if request.user.role != 'supervisor':
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})
    
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'No tienes permiso para este módulo'})
    
    # Obtener registros de entrada y salida
    registros = RegistroEntrada.objects.filter(
        trabajador__vigencia_plan=vigencia_plan
    ).select_related('trabajador')
    
    registros_data = []
    for registro in registros:
        if registro.latitud and registro.longitud:
            registros_data.append({
                'lat': float(registro.latitud),
                'lng': float(registro.longitud),
                'title': f"{registro.trabajador.get_full_name()} - Entrada - {registro.hora_entrada.strftime('%d/%m/%Y %H:%M')}",
                'type': 'entrada',
                'fecha': registro.hora_entrada.strftime('%Y-%m-%d'),
                'trabajador': registro.trabajador.id
            })
        if registro.hora_salida and registro.latitud_salida and registro.longitud_salida:
            registros_data.append({
                'lat': float(registro.latitud_salida),
                'lng': float(registro.longitud_salida),
                'title': f"{registro.trabajador.get_full_name()} - Salida - {registro.hora_salida.strftime('%d/%m/%Y %H:%M')}",
                'type': 'salida',
                'fecha': registro.hora_salida.strftime('%Y-%m-%d'),
                'trabajador': registro.trabajador.id
            })
    
    # Obtener lista de usuarios para el filtro
    usuarios = Usuario.objects.filter(vigencia_plan=vigencia_plan)
    
    context = {
        'registros_data': json.dumps(registros_data),
        'usuarios': usuarios,
        'vigencia_plan': vigencia_plan,
        'google_maps_api_key': settings.API_KEY,
        'google_maps_map_id': settings.MAP_ID
    }
    return render(request, 'Supervisores/Modulo_asistencia/maps/mapa_registros.html', context)







class ValidationView(View):
    def get(self, request):
        rut = request.GET.get('rut')
        email = request.GET.get('email')
        
        if rut:
            exists = Usuario.objects.filter(rut=rut).exists()
            return JsonResponse({'exists': exists})
            
        if email:
            exists = Usuario.objects.filter(email=email).exists()
            return JsonResponse({'exists': exists})
            
        return JsonResponse({'error': 'Campo inválido'}, status=400)
    



class GetFormTemplateView(View):
    def get(self, request, action):
        form = UsuarioForm()
        if action == 'create':
            return render(request, 'formularios/supervisor/supervisor.asistencia.html', {'form': form})
        elif action == 'edit':
            form.fields['password'].required = False  # Hacer opcional el campo password
            return render(request, 'formularios/supervisor/supervisor.edit.html', {'form': form})
        return HttpResponse(status=404)
    





@login_required
def registros_entrada_vigencia(request, vigencia_plan_id):
    if request.user.role != 'supervisor':
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})
    
    vigencia_plan = get_object_or_404(VigenciaPlan, id=vigencia_plan_id)
    if request.user.vigencia_plan != vigencia_plan:
        return render(request, 'error/error.html', {'message': 'No tienes permiso para este módulo'})
    
    registros = RegistroEntrada.objects.filter(
        trabajador__vigencia_plan=vigencia_plan
    ).select_related('trabajador').order_by('-hora_entrada')
    
    empresa_id = vigencia_plan.empresa.id
    
    context = {
        'vigencia_plan': vigencia_plan,
        'registros': registros,
        'empresa_id': empresa_id,
        'vigencia_plan_id': vigencia_plan_id,
    }
    return render(request, 'Supervisores/Modulo_asistencia/registros_entrada/registros_entrada.html', context)





# AQUI EN ADELANTE ESTA TODO LO NUEVO LA IMPLEMETNACION DE GESTION DE HORAIOS Y TURNOS Y BLOQUEOS ACCESO 

# Manejo de Horarios,creación y edición

class HorarioListView(LoginRequiredMixin, ListView):
    model = Horario
    template_name = 'Supervisores/Modulo_asistencia/horario/horarios_list.html'
    context_object_name = 'horarios'

    def get_queryset(self):
        # Filtra los horarios por la empresa del usuario autenticado
        return Horario.objects.filter(empresa=self.request.user.empresa)

    def get_context_data(self, **kwargs):
        # Obtiene el contexto base
        context = super().get_context_data(**kwargs)
        # Agrega empresa_id y vigencia_plan_id al contexto
        context['empresa_id'] = self.request.user.empresa.id if self.request.user.empresa else None
        context['vigencia_plan_id'] = self.request.user.vigencia_plan.id if self.request.user.vigencia_plan else None
        return context
class HorarioCreateView(LoginRequiredMixin, CreateView):
    model = Horario
    form_class = HorarioForm
    template_name = 'Supervisores/Modulo_asistencia/horario/horario_form.html'
    success_url = reverse_lazy('horarios_list')

    def form_valid(self, form):
        form.instance.empresa = self.request.user.empresa
        return super().form_valid(form)
    
class HorarioUpdateView(LoginRequiredMixin, UpdateView):
    model = Horario
    form_class = HorarioForm
    template_name = 'Supervisores/Modulo_asistencia/horario/horario_form.html'
    success_url = reverse_lazy('horarios_list')

class HorarioDeleteView(LoginRequiredMixin, DeleteView):
    model = Horario
    template_name = 'bases/modulo_asistencia/base_delete.html'
    success_url = reverse_lazy('horarios_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_type'] = 'horario'
        context['cancel_url'] = 'horarios_list'  # Nombre del patrón de URL para "Cancelar"
        return context










# Manejo de Turnos,Creación y Edición

class TurnoListView(LoginRequiredMixin, ListView):
    model = Turno
    template_name = 'Supervisores/Modulo_asistencia/turno/turnos_list.html'
    context_object_name = 'turnos'

    def get_queryset(self):
        return Turno.objects.filter(empresa=self.request.user.empresa)
    
    # ESTO MANEJA EL VOLVER AL HOME
    def get_queryset(self):
        # Filtra los horarios por la empresa del usuario autenticado
        return Turno.objects.filter(empresa=self.request.user.empresa)

    def get_context_data(self, **kwargs):
        # Obtiene el contexto base
        context = super().get_context_data(**kwargs)
        # Agrega empresa_id y vigencia_plan_id al contexto
        context['empresa_id'] = self.request.user.empresa.id if self.request.user.empresa else None
        context['vigencia_plan_id'] = self.request.user.vigencia_plan.id if self.request.user.vigencia_plan else None
        return context


class TurnoCreateView(LoginRequiredMixin, CreateView):
    model = Turno
    form_class = TurnoForm
    template_name = 'Supervisores/Modulo_asistencia/turno/turno_form.html'
    success_url = reverse_lazy('turnos_list')

    def form_valid(self, form):
        form.instance.empresa = self.request.user.empresa
        return super().form_valid(form)

class TurnoUpdateView(LoginRequiredMixin, UpdateView):
    model = Turno
    form_class = TurnoForm
    template_name = 'Supervisores/Modulo_asistencia/turno/turno_form.html'
    success_url = reverse_lazy('turnos_list')

class TurnoDeleteView(LoginRequiredMixin, DeleteView):
    model = Turno
    template_name = 'bases/modulo_asistencia/base_delete.html'
    success_url = reverse_lazy('turnos_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_type'] = 'Turno'
        context['cancel_url'] = 'turnos_list'  
        return context





#GENERACIÓN DE ASIGNACIONES DIARIAS

class GenerarAsignacionesView(LoginRequiredMixin, View):
    template_name = 'Supervisores/Modulo_asistencia/turno/generar_asignaciones.html'

    def get(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        form = GenerarAsignacionesForm(empresa=usuario.empresa)
        context = {
            'usuario': usuario,
            'form': form,
            'empresa_id': usuario.vigencia_plan.empresa.id,
            'vigencia_plan_id': usuario.vigencia_plan.id,
        }
        return render(request, self.template_name, context)

    def post(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        form = GenerarAsignacionesForm(request.POST, empresa=usuario.empresa)
        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            tipo_rotacion = form.cleaned_data['tipo_rotacion']
            horarios = form.cleaned_data['horarios']
            current = fecha_inicio

            if tipo_rotacion == '3_turnos_8h':
                i = 0
                while current <= fecha_fin:
                    horario = horarios[i % len(horarios)] if horarios else None
                    if horario:
                        AsignacionDiaria.objects.update_or_create(
                            usuario=usuario, fecha=current, defaults={'horario': horario}
                        )
                    i += 1
                    current += timedelta(days=1)

            elif tipo_rotacion == '12x36':
                while current <= fecha_fin:
                    horario = horarios[0] if horarios else None
                    if horario:
                        AsignacionDiaria.objects.update_or_create(
                            usuario=usuario, fecha=current, defaults={'horario': horario}
                        )
                    current += timedelta(days=2)

            elif 'x' in tipo_rotacion:
                X, Y = map(int, tipo_rotacion.split('x'))
                indice_horario = 0
                while current <= fecha_fin:
                    for _ in range(X):
                        if current > fecha_fin:
                            break
                        horario = horarios[indice_horario % len(horarios)]
                        AsignacionDiaria.objects.update_or_create(
                            usuario=usuario, fecha=current, defaults={'horario': horario}
                        )
                        indice_horario += 1
                        current += timedelta(days=1)
                    for _ in range(Y):
                        if current > fecha_fin:
                            break
                        AsignacionDiaria.objects.filter(usuario=usuario, fecha=current).delete()
                        current += timedelta(days=1)

            elif tipo_rotacion == 'personalizado':
                X = form.cleaned_data['dias_trabajo']
                Y = form.cleaned_data['dias_descanso']
                indice_horario = 0
                while current <= fecha_fin:
                    for _ in range(X):
                        if current > fecha_fin:
                            break
                        horario = horarios[indice_horario % len(horarios)]
                        AsignacionDiaria.objects.update_or_create(
                            usuario=usuario, fecha=current, defaults={'horario': horario}
                        )
                        indice_horario += 1
                        current += timedelta(days=1)
                    for _ in range(Y):
                        if current > fecha_fin:
                            break
                        AsignacionDiaria.objects.filter(usuario=usuario, fecha=current).delete()
                        current += timedelta(days=1)

            messages.success(request, f'Asignaciones generadas para {usuario.get_full_name()}.')
            return redirect('calendario_turno', user_id=user_id)
        else:
            context = {
                'usuario': usuario,
                'form': form,
                'empresa_id': usuario.vigencia_plan.empresa.id,
                'vigencia_plan_id': usuario.vigencia_plan.id,
            }
            return render(request, self.template_name, context)






    
# Creación y edición de usuarios

class UserCreateUpdateView(LoginRequiredMixin, View):
    def get(self, request, vigencia_plan_id, user_id=None):  
        if user_id:
            user = get_object_or_404(Usuario, pk=user_id)
            data = {
                'username': user.username,
                'rut': user.rut,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'celular': user.celular, 
                'role': user.role,
                'horario': user.horario_id,
                'turno': user.turno_id,
                'metodo_registro_permitido': user.metodo_registro_permitido
            }
            return JsonResponse(data)
        return JsonResponse({'error': 'ID de usuario no proporcionado'}, status=400)

    def delete(self, request, vigencia_plan_id, user_id):
        user = get_object_or_404(Usuario, pk=user_id)
        try:
            user.delete()
            return JsonResponse({'message': 'Usuario eliminado exitosamente'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def post(self, request, vigencia_plan_id, user_id=None):
        vigencia_plan = get_object_or_404(VigenciaPlan, pk=vigencia_plan_id)
        user = get_object_or_404(Usuario, pk=user_id) if user_id else None
        
        form = UsuarioForm(request.POST, instance=user)
        if form.is_valid():
            if form.cleaned_data['role'] == 'admin':
                return JsonResponse({'errors': {'role': ['No se permite asignar el rol de administrador']}}, status=403)
            
            user = form.save(commit=False)
            user.empresa = vigencia_plan.empresa
            user.vigencia_plan = vigencia_plan
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()
            return JsonResponse({
                'message': 'Usuario guardado exitosamente',
                'redirect_url': reverse('supervisor_home_asistencia', kwargs={
                    'empresa_id': vigencia_plan.empresa.id,
                    'vigencia_plan_id': vigencia_plan.id
                })
            }, status=200)
        else:
            errors = {f: [str(e) for e in e_list] for f, e_list in form.errors.items()}
            return JsonResponse({'errors': errors}, status=400)
        









# Manejo de Calendario de Turnos
class CalendarioTurnoView(LoginRequiredMixin, View):
    template_name = 'Supervisores/Modulo_asistencia/turno/calendario_turno.html'

    def get(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        año = int(request.GET.get('año', date.today().year))
        horarios = Horario.objects.filter(empresa=usuario.empresa)

        # Generar todos los días del año
        inicio_año = date(año, 1, 1)
        fin_año = date(año, 12, 31)
        dias = []
        current = inicio_año
        while current <= fin_año:
            asignacion = AsignacionDiaria.objects.filter(usuario=usuario, fecha=current).first()
            dias.append({
                'fecha': current,
                'horario': asignacion.horario if asignacion else None,
                'mes': current.month,
                'dia_semana': current.weekday(),
            })
            current += timedelta(days=1)

        # Organizar por meses
        meses = {i: [d for d in dias if d['mes'] == i] for i in range(1, 13)}

        context = {
            'usuario': usuario,
            'año': año,
            'meses': meses,
            'horarios': horarios,
            'empresa_id': usuario.vigencia_plan.empresa.id,
            'vigencia_plan_id': usuario.vigencia_plan.id,
        }
        return render(request, self.template_name, context)

    def post(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        año = int(request.POST.get('año', date.today().year))
        horarios_enviados = {k: v for k, v in request.POST.items() if k.startswith('horario_')}

        for key, horario_id in horarios_enviados.items():
            fecha = date.fromisoformat(key.split('_')[1])
            if horario_id:
                horario = Horario.objects.get(id=horario_id)
                AsignacionDiaria.objects.update_or_create(
                    usuario=usuario, fecha=fecha, defaults={'horario': horario}
                )
            else:
                AsignacionDiaria.objects.filter(usuario=usuario, fecha=fecha).delete()

        messages.success(request, f'Horarios de {usuario.get_full_name()} ({año}) guardados.')
        return redirect('supervisor_home_asistencia', 
                        empresa_id=usuario.vigencia_plan.empresa.id, 
                        vigencia_plan_id=usuario.vigencia_plan.id)
    





class ActualizarDiaView(LoginRequiredMixin, View):
    def post(self, request, user_id):
        usuario = get_object_or_404(Usuario, pk=user_id)
        fecha = request.POST.get('fecha')
        horario_id = request.POST.get('horario_id')

        if not fecha:
            return JsonResponse({'success': False, 'error': 'Fecha no proporcionada'}, status=400)

        if horario_id:
            horario = Horario.objects.get(id=horario_id)
            AsignacionDiaria.objects.update_or_create(
                usuario=usuario, fecha=fecha, defaults={'horario': horario}
            )
        else:
            AsignacionDiaria.objects.filter(usuario=usuario, fecha=fecha).delete()

        return JsonResponse({'success': True})
    





# Función para exportar querysets a Excel
def export_queryset_to_excel(request, queryset, fields, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(fields)  # Añadir encabezados
    for item in queryset:
        data = [str(getattr(item, field, '')) for field in fields]
        ws.append(data)
    if not queryset.exists():
        ws.append(["No hay datos registrados."])
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    response = HttpResponse(
        excel_buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response












@login_required
def user_full_info(request, user_id):
    user = get_object_or_404(Usuario, id=user_id)
    
    # Verificación de permisos
    if request.user.role not in ['supervisor', 'admin'] or user.vigencia_plan != request.user.vigencia_plan:
        return render(request, 'error/error.html', {'message': 'Acceso no autorizado'})

    # Obtener o crear instancias de modelos OneToOne
    perfil, _ = PerfilUsuario.objects.get_or_create(usuario=user)
    contacto, _ = ContactoUsuario.objects.get_or_create(usuario=user)
    info_bancaria, _ = InformacionBancaria.objects.get_or_create(usuario=user)
    info_adicional, _ = InformacionAdicional.objects.get_or_create(usuario=user)
    seguro_cesantia, _ = SeguroCesantia.objects.get_or_create(usuario=user)
    prevision, _ = Prevision.objects.get_or_create(usuario=user)
    otros, _ = Otros.objects.get_or_create(usuario=user)
    info_complementaria, _ = InformacionComplementaria.objects.get_or_create(usuario=user)

    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        # Manejo de formularios OneToOne con depuración
        if form_type == 'usuario':
            form = UsuarioForm(request.POST, instance=user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información básica actualizada.')
            else:
                messages.error(request, f'Error en información básica: {form.errors}')

        elif form_type == 'perfil':
            form = PerfilUsuarioForm(request.POST, instance=perfil)
            if form.is_valid():
                form.save()
                messages.success(request, 'Perfil actualizado.')
            else:
                messages.error(request, f'Error en perfil: {form.errors}')

        elif form_type == 'contacto':
            form = ContactoUsuarioForm(request.POST, instance=contacto)
            if form.is_valid():
                form.save()
                messages.success(request, 'Contacto actualizado.')
            else:
                messages.error(request, f'Error en contacto: {form.errors}')

        elif form_type == 'info_bancaria':
            form = InformacionBancariaForm(request.POST, instance=info_bancaria)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información bancaria actualizada.')
            else:
                messages.error(request, f'Error en info bancaria: {form.errors}')

        elif form_type == 'info_adicional':
            form = InformacionAdicionalForm(request.POST, instance=info_adicional)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información adicional actualizada.')
            else:
                messages.error(request, f'Error en info adicional: {form.errors}')

        elif form_type == 'seguro_cesantia':
            form = SeguroCesantiaForm(request.POST, instance=seguro_cesantia)
            if form.is_valid():
                form.save()
                messages.success(request, 'Seguro de cesantía actualizado.')
            else:
                messages.error(request, f'Error en seguro de cesantía: {form.errors}')

        elif form_type == 'prevision':
            form = PrevisionForm(request.POST, instance=prevision)
            if form.is_valid():
                form.save()
                messages.success(request, 'Previsión actualizada.')
            else:
                messages.error(request, f'Error en previsión: {form.errors}')

        elif form_type == 'otros':
            form = OtrosForm(request.POST, instance=otros)
            if form.is_valid():
                form.save()
                messages.success(request, 'Otros datos actualizados.')
            else:
                messages.error(request, f'Error en otros: {form.errors}')

        elif form_type == 'info_complementaria':
            form = InformacionComplementariaForm(request.POST, instance=info_complementaria)
            if form.is_valid():
                form.save()
                messages.success(request, 'Información complementaria actualizada.')
            else:
                messages.error(request, f'Error en info complementaria: {form.errors}')

        # Manejo de formsets ForeignKey con depuración
        elif form_type == 'antecedentes_conducir':
            formset = AntecedentesConducirFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Antecedentes de conducir actualizados.')
            else:
                messages.error(request, f'Error en antecedentes de conducir: {formset.errors}')
            if 'export_antecedentes_conducir' in request.POST:
                queryset = user.antecedentes_conducir.all()
                fields = ['tipo_licencia', 'municipalidad', 'fecha_ultimo_control', 
                          'fecha_vencimiento', 'hoja_vida_conducir']
                return export_queryset_to_excel(request, queryset, fields, f"antecedentes_conducir_{user.username}")

        elif form_type == 'nivel_estudios':
            formset = NivelEstudiosFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Niveles de estudios actualizados.')
            else:
                messages.error(request, f'Error en nivel de estudios: {formset.errors}')
            if 'export_nivel_estudios' in request.POST:
                queryset = user.niveles_estudios.all()
                fields = ['nivel_estudios', 'completo', 'ultimo_curso', 'carrera']
                return export_queryset_to_excel(request, queryset, fields, f"nivel_estudios_{user.username}")

        elif form_type == 'examenes':
            formset = ExamenesMutualFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Exámenes mutual actualizados.')
            else:
                messages.error(request, f'Error en exámenes: {formset.errors}')
            if 'export_examenes' in request.POST:
                queryset = user.examenes_mutual.all()
                fields = ['tipo_examen', 'fecha_examen', 'fecha_vencimiento']
                return export_queryset_to_excel(request, queryset, fields, f"examenes_mutual_{user.username}")

        elif form_type == 'grupo_familiar':
            formset = GrupoFamiliarFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Grupo familiar actualizado.')
            else:
                messages.error(request, f'Error en grupo familiar: {formset.errors}')
            if 'export_grupo_familiar' in request.POST:
                queryset = user.grupo_familiar.all()
                fields = ['rut_carga', 'nombre_carga', 'fecha_nacimiento', 'edad', 'sexo']
                return export_queryset_to_excel(request, queryset, fields, f"grupo_familiar_{user.username}")

        elif form_type == 'capacitacion':
            formset = CapacitacionFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Capacitaciones actualizadas.')
            else:
                messages.error(request, f'Error en capacitaciones: {formset.errors}')
            if 'export_capacitaciones' in request.POST:
                queryset = user.capacitaciones.all()
                fields = ['descripcion', 'horas', 'institucion']
                return export_queryset_to_excel(request, queryset, fields, f"capacitaciones_{user.username}")

        elif form_type == 'licencias_medicas':
            formset = LicenciasMedicasFormSet(request.POST, instance=user)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Licencias médicas actualizadas.')
            else:
                messages.error(request, f'Error en licencias médicas: {formset.errors}')
            if 'export_licencias_medicas' in request.POST:
                queryset = user.licencias_medicas.all()
                fields = ['tipo_accidente', 'clasificacion_accidente', 'fecha_inicio_reposo', 
                          'fecha_termino', 'fecha_alta', 'dias_reposo']
                return export_queryset_to_excel(request, queryset, fields, f"licencias_medicas_{user.username}")

        return redirect('user_full_info', user_id=user.id)

    # GET: Preparar formularios y formsets
    usuario_form = UsuarioForm(instance=user)
    perfil_form = PerfilUsuarioForm(instance=perfil)
    contacto_form = ContactoUsuarioForm(instance=contacto)
    info_bancaria_form = InformacionBancariaForm(instance=info_bancaria)
    info_adicional_form = InformacionAdicionalForm(instance=info_adicional)
    seguro_cesantia_form = SeguroCesantiaForm(instance=seguro_cesantia)
    prevision_form = PrevisionForm(instance=prevision)
    otros_form = OtrosForm(instance=otros)
    info_complementaria_form = InformacionComplementariaForm(instance=info_complementaria)
    
    antecedentes_conducir_formset = AntecedentesConducirFormSet(instance=user)
    nivel_estudios_formset = NivelEstudiosFormSet(instance=user)
    examenes_formset = ExamenesMutualFormSet(instance=user)
    grupo_familiar_formset = GrupoFamiliarFormSet(instance=user)
    capacitacion_formset = CapacitacionFormSet(instance=user)
    licencias_medicas_formset = LicenciasMedicasFormSet(instance=user)

    context = {
        'user': user,
        'usuario_form': usuario_form,
        'perfil_form': perfil_form,
        'contacto_form': contacto_form,
        'info_bancaria_form': info_bancaria_form,
        'info_adicional_form': info_adicional_form,
        'seguro_cesantia_form': seguro_cesantia_form,
        'prevision_form': prevision_form,
        'otros_form': otros_form,
        'info_complementaria_form': info_complementaria_form,
        'antecedentes_conducir_formset': antecedentes_conducir_formset,
        'nivel_estudios_formset': nivel_estudios_formset,
        'examenes_formset': examenes_formset,
        'grupo_familiar_formset': grupo_familiar_formset,
        'capacitacion_formset': capacitacion_formset,
        'licencias_medicas_formset': licencias_medicas_formset,
    }
    
    return render(request, 'Supervisores/Modulo_asistencia/DatosCompletos/usuario_datos.html', context)















def export_selected_tabs(request, user_id):
    user = get_object_or_404(Usuario, id=user_id)
    
    if request.method == 'POST':
        selected_tabs = request.POST.getlist('tabs')
        combine = 'combine' in request.POST
        
        if not selected_tabs:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sin Datos"
            ws.append(["No se seleccionaron pestañas para exportar."])
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            response = HttpResponse(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="export_{user.username}.xlsx"'
            return response

        if combine:
            wb = Workbook()
            wb.remove(wb.active)
            for tab in selected_tabs:
                ws = wb.create_sheet(title=tab.capitalize())
                if tab == 'usuario':
                    fields = ['rut', 'username', 'first_name', 'last_name', 'email']
                    ws.append(fields)
                    data = [str(getattr(user, field, '')) for field in fields]
                    ws.append(data if any(data) else ["No hay datos disponibles."])
                elif tab == 'perfil':
                    try:
                        perfil = PerfilUsuario.objects.get(usuario=user)
                        fields = ['apellido_paterno', 'apellido_materno', 'fecha_nacimiento', 'cargo']
                        ws.append(fields)
                        data = [str(getattr(perfil, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except PerfilUsuario.DoesNotExist:
                        ws.append(["No se encontró el perfil del usuario."])
                elif tab == 'contacto':
                    try:
                        contacto = ContactoUsuario.objects.get(usuario=user)
                        fields = ['direccion', 'numero', 'telefono', 'celular', 'region', 'provincia', 'comuna']
                        ws.append(fields)
                        data = [str(getattr(contacto, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except ContactoUsuario.DoesNotExist:
                        ws.append(["No se encontró información de contacto."])
                elif tab == 'info_bancaria':
                    try:
                        info_bancaria = InformacionBancaria.objects.get(usuario=user)
                        fields = ['banco', 'tipo_cuenta', 'numero_cuenta']
                        ws.append(fields)
                        data = [str(getattr(info_bancaria, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except InformacionBancaria.DoesNotExist:
                        ws.append(["No se encontró información bancaria."])
                elif tab == 'info_adicional':
                    try:
                        info_adicional = InformacionAdicional.objects.get(usuario=user)
                        fields = ['fecha_primera_cotizacion', 'meses_anteriores', 'anos_anteriores']
                        ws.append(fields)
                        data = [str(getattr(info_adicional, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except InformacionAdicional.DoesNotExist:
                        ws.append(["No se encontró información adicional."])
                elif tab == 'seguro_cesantia':
                    try:
                        seguro_cesantia = SeguroCesantia.objects.get(usuario=user)
                        fields = ['acogido_seguro', 'afp_recaudadora', 'sueldo_patronal']
                        ws.append(fields)
                        data = [str(getattr(seguro_cesantia, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except SeguroCesantia.DoesNotExist:
                        ws.append(["No se encontró información de seguro de cesantía."])
                elif tab == 'prevision':
                    try:
                        prevision = Prevision.objects.get(usuario=user)
                        fields = ['salud', 'tasa', 'regimen', 'afp']
                        ws.append(fields)
                        data = [str(getattr(prevision, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except Prevision.DoesNotExist:
                        ws.append(["No se encontró información de previsión."])
                elif tab == 'otros':
                    try:
                        otros = Otros.objects.get(usuario=user)
                        fields = ['tipo_discapacidad', 'tasa_indemnizacion']
                        ws.append(fields)
                        data = [str(getattr(otros, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except Otros.DoesNotExist:
                        ws.append(["No se encontró información adicional."])
                elif tab == 'nivel_estudios':
                    try:
                        nivel_estudios = NivelEstudios.objects.get(usuario=user)
                        fields = ['nivel_estudios', 'completo', 'ultimo_curso', 'carrera']
                        ws.append(fields)
                        data = [str(getattr(nivel_estudios, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except NivelEstudios.DoesNotExist:
                        ws.append(["No se encontró nivel de estudios."])
                elif tab == 'info_complementaria':
                    try:
                        info_complementaria = InformacionComplementaria.objects.get(usuario=user)
                        fields = ['pais_origen', 'pasaporte', 'estado_civil', 'tipo_visa']
                        ws.append(fields)
                        data = [str(getattr(info_complementaria, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    except InformacionComplementaria.DoesNotExist:
                        ws.append(["No se encontró información complementaria."])
                elif tab == 'examenes':
                    queryset = user.examenes_mutual.all()
                    if queryset.exists():
                        fields = ['tipo_examen', 'fecha_examen', 'fecha_vencimiento']
                        ws.append(fields)
                        for item in queryset:
                            data = [str(getattr(item, field, '')) for field in fields]
                            ws.append(data)
                    else:
                        ws.append(["No hay exámenes registrados."])
                elif tab == 'grupo_familiar':
                    queryset = user.grupo_familiar.all()
                    if queryset.exists():
                        fields = ['rut_carga', 'nombre_carga', 'fecha_nacimiento', 'edad', 'sexo']
                        ws.append(fields)
                        for item in queryset:
                            data = [str(getattr(item, field, '')) for field in fields]
                            ws.append(data)
                    else:
                        ws.append(["No hay grupo familiar registrado."])
                elif tab == 'capacitacion':
                    queryset = user.capacitaciones.all()
                    if queryset.exists():
                        fields = ['descripcion', 'horas', 'institucion']
                        ws.append(fields)
                        for item in queryset:
                            data = [str(getattr(item, field, '')) for field in fields]
                            ws.append(data)
                    else:
                        ws.append(["No hay capacitaciones registradas."])
                elif tab == 'licencias_medicas':
                    queryset = user.licencias_medicas.all()
                    if queryset.exists():
                        fields = ['tipo_accidente', 'clasificacion_accidente', 'fecha_inicio_reposo', 
                                  'fecha_termino', 'fecha_alta', 'dias_reposo']
                        ws.append(fields)
                        for item in queryset:
                            data = [str(getattr(item, field, '')) for field in fields]
                            ws.append(data)
                    else:
                        ws.append(["No hay licencias médicas registradas."])
                elif tab == 'antecedentes_conducir':
                    queryset = user.antecedentes_conducir.all()
                    if queryset.exists():
                        fields = ['tipo_licencia', 'municipalidad', 'fecha_ultimo_control', 
                                  'fecha_vencimiento', 'hoja_vida_conducir']
                        ws.append(fields)
                        for item in queryset:
                            data = [str(getattr(item, field, '')) for field in fields]
                            ws.append(data)
                    else:
                        ws.append(["No hay antecedentes de conducir registrados."])

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            response = HttpResponse(
                excel_buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="export_{user.username}.xlsx"'
            return response
        else:
            from zipfile import ZipFile
            zip_buffer = BytesIO()
            with ZipFile(zip_buffer, 'w') as zip_file:
                for tab in selected_tabs:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = tab.capitalize()
                    if tab == 'usuario':
                        fields = ['rut', 'username', 'first_name', 'last_name', 'email']
                        ws.append(fields)
                        data = [str(getattr(user, field, '')) for field in fields]
                        ws.append(data if any(data) else ["No hay datos disponibles."])
                    elif tab == 'perfil':
                        try:
                            perfil = PerfilUsuario.objects.get(usuario=user)
                            fields = ['apellido_paterno', 'apellido_materno', 'fecha_nacimiento', 'cargo']
                            ws.append(fields)
                            data = [str(getattr(perfil, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except PerfilUsuario.DoesNotExist:
                            ws.append(["No se encontró el perfil del usuario."])
                    elif tab == 'contacto':
                        try:
                            contacto = ContactoUsuario.objects.get(usuario=user)
                            fields = ['direccion', 'numero', 'telefono', 'celular', 'region', 'provincia', 'comuna']
                            ws.append(fields)
                            data = [str(getattr(contacto, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except ContactoUsuario.DoesNotExist:
                            ws.append(["No se encontró información de contacto."])
                    elif tab == 'info_bancaria':
                        try:
                            info_bancaria = InformacionBancaria.objects.get(usuario=user)
                            fields = ['banco', 'tipo_cuenta', 'numero_cuenta']
                            ws.append(fields)
                            data = [str(getattr(info_bancaria, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except InformacionBancaria.DoesNotExist:
                            ws.append(["No se encontró información bancaria."])
                    elif tab == 'info_adicional':
                        try:
                            info_adicional = InformacionAdicional.objects.get(usuario=user)
                            fields = ['fecha_primera_cotizacion', 'meses_anteriores', 'anos_anteriores']
                            ws.append(fields)
                            data = [str(getattr(info_adicional, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except InformacionAdicional.DoesNotExist:
                            ws.append(["No se encontró información adicional."])
                    elif tab == 'seguro_cesantia':
                        try:
                            seguro_cesantia = SeguroCesantia.objects.get(usuario=user)
                            fields = ['acogido_seguro', 'afp_recaudadora', 'sueldo_patronal']
                            ws.append(fields)
                            data = [str(getattr(seguro_cesantia, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except SeguroCesantia.DoesNotExist:
                            ws.append(["No se encontró información de seguro de cesantía."])
                    elif tab == 'prevision':
                        try:
                            prevision = Prevision.objects.get(usuario=user)
                            fields = ['salud', 'tasa', 'regimen', 'afp']
                            ws.append(fields)
                            data = [str(getattr(prevision, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except Prevision.DoesNotExist:
                            ws.append(["No se encontró información de previsión."])
                    elif tab == 'otros':
                        try:
                            otros = Otros.objects.get(usuario=user)
                            fields = ['tipo_discapacidad', 'tasa_indemnizacion']
                            ws.append(fields)
                            data = [str(getattr(otros, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except Otros.DoesNotExist:
                            ws.append(["No se encontró información adicional."])
                    elif tab == 'nivel_estudios':
                        try:
                            nivel_estudios = NivelEstudios.objects.get(usuario=user)
                            fields = ['nivel_estudios', 'completo', 'ultimo_curso', 'carrera']
                            ws.append(fields)
                            data = [str(getattr(nivel_estudios, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except NivelEstudios.DoesNotExist:
                            ws.append(["No se encontró nivel de estudios."])
                    elif tab == 'info_complementaria':
                        try:
                            info_complementaria = InformacionComplementaria.objects.get(usuario=user)
                            fields = ['pais_origen', 'pasaporte', 'estado_civil', 'tipo_visa']
                            ws.append(fields)
                            data = [str(getattr(info_complementaria, field, '')) for field in fields]
                            ws.append(data if any(data) else ["No hay datos disponibles."])
                        except InformacionComplementaria.DoesNotExist:
                            ws.append(["No se encontró información complementaria."])
                    elif tab == 'examenes':
                        queryset = user.examenes_mutual.all()
                        if queryset.exists():
                            fields = ['tipo_examen', 'fecha_examen', 'fecha_vencimiento']
                            ws.append(fields)
                            for item in queryset:
                                data = [str(getattr(item, field, '')) for field in fields]
                                ws.append(data)
                        else:
                            ws.append(["No hay exámenes registrados."])
                    elif tab == 'grupo_familiar':
                        queryset = user.grupo_familiar.all()
                        if queryset.exists():
                            fields = ['rut_carga', 'nombre_carga', 'fecha_nacimiento', 'edad', 'sexo']
                            ws.append(fields)
                            for item in queryset:
                                data = [str(getattr(item, field, '')) for field in fields]
                                ws.append(data)
                        else:
                            ws.append(["No hay grupo familiar registrado."])
                    elif tab == 'capacitacion':
                        queryset = user.capacitaciones.all()
                        if queryset.exists():
                            fields = ['descripcion', 'horas', 'institucion']
                            ws.append(fields)
                            for item in queryset:
                                data = [str(getattr(item, field, '')) for field in fields]
                                ws.append(data)
                        else:
                            ws.append(["No hay capacitaciones registradas."])
                    elif tab == 'licencias_medicas':
                        queryset = user.licencias_medicas.all()
                        if queryset.exists():
                            fields = ['tipo_accidente', 'clasificacion_accidente', 'fecha_inicio_reposo', 
                                      'fecha_termino', 'fecha_alta', 'dias_reposo']
                            ws.append(fields)
                            for item in queryset:
                                data = [str(getattr(item, field, '')) for field in fields]
                                ws.append(data)
                        else:
                            ws.append(["No hay licencias médicas registradas."])
                    elif tab == 'antecedentes_conducir':
                        queryset = user.antecedentes_conducir.all()
                        if queryset.exists():
                            fields = ['tipo_licencia', 'municipalidad', 'fecha_ultimo_control', 
                                      'fecha_vencimiento', 'hoja_vida_conducir']
                            ws.append(fields)
                            for item in queryset:
                                data = [str(getattr(item, field, '')) for field in fields]
                                ws.append(data)
                        else:
                            ws.append(["No hay antecedentes de conducir registrados."])
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    zip_file.writestr(f"{tab}_{user.username}.xlsx", excel_buffer.getvalue())
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="export_{user.username}.zip"'
            return response
    else:
        return HttpResponse("Método no permitido", status=405)